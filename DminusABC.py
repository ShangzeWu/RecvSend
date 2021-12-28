# -*- coding: UTF-8 -*-
from openpyxl import *
import os
import time
from datetime import datetime, timedelta
import pandas as pd
import xlrd

format_pattern = '%Y-%m-%d %H:%M:%S'
cur_time = datetime.now()
# 将 'cur_time' 类型时间通过格式化模式转换为 'str' 时间
cur_time = cur_time.strftime(format_pattern)

def find_new_file(dir):
    #查找目录下最新的文件
    file_lists = os.listdir(dir)
    file_lists.sort(key=lambda fn: os.path.getmtime(dir + "/" + fn)
    if not os.path.isdir(dir + "/" + fn) else 0)
#    print('最新的文件为： ' + file_lists[-1])
    file = os.path.join(dir, file_lists[-1])
#    print('完整路径：', file)
    return file_lists[-1]   #返回文件的名字，不包含路径

path =  "/var/www/html/RecvSend/"
#print(path)
dir_A = path+'/uploadA/' #用来读取A文件 的 路径
dir_B = path+'/uploadB/' #用来读取B文件 的 路径
dir_C = path+'/uploadC/' #用来读取C文件 的 路径
dir_D = path+'/uploadD/' #用来读取D文件 的 路径

dir_save_D= "/var/www/html/RecvSend/resultD/"  #输出 D文件 的保存路径

file_name_A = find_new_file(dir_A)
file_name_B = find_new_file(dir_B)
file_name_C = find_new_file(dir_C)
file_name_D = find_new_file(dir_D)

#业务逻辑
#加载ABC表的第一列
wb1 = load_workbook(dir_A+file_name_A) #A表
ws1 = wb1[wb1.sheetnames[0]]           #A表第一页
wb2 = load_workbook(dir_B+file_name_B) #B表
ws2 = wb2[wb2.sheetnames[0]]           #B表第一页
wb3 = load_workbook(dir_C+file_name_C) #C表
ws3 = wb3[wb3.sheetnames[0]]           #C表第一页

Allrow1 = ws1.max_row
#Allcol1 = ws1.max_column
Allrow2 = ws2.max_row
Allrow3 = ws3.max_row

list_number = [0]
#print(type(ws1.cell(3,1).value))  #int
#print(type(ws2.cell(3,1).value))  #int
#print(type(ws3.cell(3,1).value))  #int

#使用pandas剔除杂项
dfA = pd.read_excel(dir_A+file_name_A)
dfB = pd.read_excel(dir_B+file_name_B)
dfC = pd.read_excel(dir_C+file_name_C)

ListA = dfA['运单号'].values.tolist()
ListB = dfB['运单号'].values.tolist()
ListC = dfC['运单号'].values.tolist()

list_number.extend(ListA)
list_number.extend(ListB)
list_number.extend(ListC)

print(list_number)
#print(df['寄件网点'])
#print((df['寄件网点']!='江苏省市场部五十七部') & (df['寄件网点']!='江苏盐城公司'))
df = pd.read_excel(dir_D+file_name_D)
#删除其他网点
df = df.drop(df[(df['寄件网点']!='江苏省市场部五十七部') & (df['寄件网点']!='江苏盐城公司') & (df['寄件网点']!='江苏盐城宝龙公司') & (df['寄件网点']!='江苏盐城龙冈公司') & (df['寄件网点']!='江苏盐城亭湖公司') & (df['寄件网点']!='江苏盐城万达公司') & (df['寄件网点']!='江苏盐城吾悦公司') & (df['寄件网点']!='江苏盐城盐都公司') & (df['寄件网点']!='江苏盐城盐南高新公司') & (df['寄件网点']!='江苏盐城招商公司')  ].index)
#删除空行
df = df.dropna(axis=0, how='all', thresh=None, subset=None, inplace=False)
# print(df['运单编号'].dtype) #int64
df = df.drop_duplicates(subset='运单编号', keep='first', inplace=False)

for num1 in list_number:
    num1 = int(num1)
    #print(type(num1))
    #print(df['运单编号'].dtype)
    df = df.drop(df[ df['运单编号'] == num1 ].index)

#df = df.drop(df[ df['运单编号'] == 777069457504657].index)
writer = pd.ExcelWriter(path+'/resultD/'+cur_time+'.xlsx')
#df为需要保存的DataFrame
df.to_excel(writer,index = False ,encoding='utf-8',sheet_name='Sheet1')
writer.save()
#wb2.save(dir_namelist+file_name_list)
